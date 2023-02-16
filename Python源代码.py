import os
from openpyxl import Workbook
a=["Day","Np237","Am241","Am243","Cm244"]
V = [29.82894964,44.74342446,59.65789928,74.5723741]
F = [0.56363,0.26477,0.12035,0.05125] #MA各核素对应含量
rou = 10.4  # MA材料密度
Q = [0.1,0.2,0.5,1,2,5,10]
Na = 6.02e+23 #阿伏伽德罗常数
M = [237,241,243,244] #相对原子质量
nrou = 1e+24 #原子密度10^24/cm3
N = 1536 #整个反应堆放置了1536根嬗变棒
word=['A','B','C','D']
_word = ['1','2','3','4','5','6','7']  #依据文件命名区分设定的方案，如A1对应MA&UO2厚度d=0.1cm,比例Q=0.1

filename ="无6LiD数据处理.xlsx"
name='den_tot'
openfile=[]
n,b=0,[]
L = len(os.path.abspath('.'))
print(L)
name='den_tot'
def searchfile(path):
    for item in os.listdir(path):
        if os.path.isdir(os.path.join(path, item)):
            searchfile(os.path.join(path, item))
        else:
            if name in item:
                print(item)
                b.append(item)
                openfile.append(path+'\\'+item)
print("处理的文件有：")
searchfile(os.path.abspath('.'))  #调用寻找文件路径函数
input("请按回车确认执行")
lsin=os.listdir()
# 寻找同目录下的'den_tot'文件
wb=Workbook()
sheet=wb.active
sheet.title="MA嬗变率统计"
#新建一个Excel表格，sheet命名为"MA嬗变率统计"

for m in range(len(openfile)):
    fo = open (openfile[m],'r')
    s = fo.readlines()
    list1=[]
    list3=[]
    for item in s:
        for i in a:
            if i in item:
                list1.append(item)
    fo.close()
#打开txt文件,提取所需数据存入list1列表中
    c=[]
    c.append(b[m]+'文件中')
    for i in range(len(word)):
        if word[i] in b[m]:
            w = i
            for j in range(len(_word)):
                if _word[j] in b[m]:
                    z = j
    #标记在哪一个方案中，A1对应体积V=29.82894964cm3,MA:UO2比值Q=0.1依次类推。
    sheet.append(c)
    sheet.append(a)
    for item in list1:
        list2 = item.split()
        for i in range(len(a)):   #依次判断a中所有元素，提取所在行的关键信息
            if a[i] in item:
                if i ==0:
                    list3.append(eval(list2[3]))
                else :
                    list3.append(eval(list2[2]))
    x = n #x用来存储前一个n值
    n = int(len(list1) / len(a))# 判断设定的燃耗步长
    for i in range(n):
        sheet.append(list3[i*5:i*5+5])
    sheet.cell(row=2, column=6).value = a[0]
    sheet.cell(row=1, column=7).value = '嬗变率分别为:'
    for i in range(n):
        sheet.cell(row=m * (x + 2) + i + 3, column=6).value = "=A{}".format(m*(x+2)+i+3)#day数据填充
        for j in range(len(a)-1):
            sheet.cell(row=m * (x + 2)+i+3, column=j+12).value="={1}{2}*{3}-{1}{0}*{3}".format(m*(x+2)+i+3,chr(j+66),m*(x+2)+3,V[w]*nrou*M[j]*N/Na)
            sheet.cell(row=m*(x+2)+i+3,column=j+7).value = "=({1}{2}-{1}{0})/{1}{2}".format(m*(x+2)+i+3,chr(j+66),m*(x+2)+3)
    for j in range(len(a)-1):
        sheet.cell(row=1, column=j+17).value = a[j+1]
        sheet.cell(row=m+2, column=j+17).value = "={1}{0}*{2}".format(m*(x+2)+n+2,chr(j+71),V[w]*rou*Q[z]*N*F[j]/(Q[z]+1))
        sheet.cell(row=m+2, column=j+22).value = "={1}{0}*{2}".format(m*(x+2)+n+2, chr(j+71),100)
    # 实现每个时长下嬗变率的计算,i为行循环次数，j为列循环次数，m为文件循环次数，x为前一次的步长循环次数（）
wb.save(filename)