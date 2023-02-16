from openpyxl import Workbook
import os
a = ["Day","O16","U235","U238","Np237","Am241","Am243","Cm244"] #所要选取的核素，可依情况更改,核素需按原子序数先后排序
v = "9.4951296,14.2426944,18.9902592,23.737824"
Q = "0.1,0.2,0.5,1,2,5,10"

filename = "数据处理.xlsx"
name='den_tot'
b=[]
lena = len(a)
openfile=[]
ls=os.getcwd()
ls=ls+"\\{}".format(fname)
lsin=os.listdir(ls)
for i in range(len(lsin)):
    ls1=ls+"\\{}".format(lsin[i])
    b.append(ls1[-2:])
    ls2=os.listdir(ls1)
    for j in ls2:
        if name in j :
            openfile.append(ls1+'\\{}'.format(j))

wb=Workbook()
sheet=wb.active
sheet.title="MA嬗变率统计"
#新建一个Excel表格，sheet命名为"MA嬗变率统计"

for m in range(len(b)):
    file = openfile[m]  #核素燃耗的文件名
    fo = open(file,'r')
    s = fo.readlines()
    list1=[]
    list3=[]
    for item in s:
        for i in a:
            if i in item:
                list1.append(item)
    fo.close()
#打开txt文件,提取所需数据存入list1列表中
    c=[]
    c.append(b[m]+'方案')
    sheet.append(c)
    sheet.append(a)
    for item in list1:
        list2 = item.split()
        for i in range(lena):   #依次判断a中所有元素，提取所在行的关键信息
            if a[i] in item:
                if i ==0:
                    list3.append(eval(list2[3]))
                else :
                    list3.append(eval(list2[2]))
    for i in range(11):
        sheet.append(list3[i*lena:i*lena+lena])

    # sheet.cell(row=1, column=6).value = '嬗变率依次为'
    # for i in range(11):
    #     for j in range(len(a)):
    #         sheet.cell(row=m*13+i+3,column=j+6).value = "=({1}{2}-{1}{0})/{1}{2}".format(m*13+i+3,chr(j+65),m*13+3)
wb.save(filename)




